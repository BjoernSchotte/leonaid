#!/usr/bin/env python3
"""Dry-run and import Company/Person rows through the semantic CRM gateway."""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import json
import os
import re
import sys
from collections import Counter
from dataclasses import asdict, dataclass, replace
from datetime import datetime
from enum import StrEnum
from pathlib import Path
from typing import Any, Literal, cast
from uuid import UUID

from openpyxl import load_workbook
from pydantic import SecretStr

from leonaid.adapters.twenty.gateway import TwentyCrmGateway, TwentyGatewaySettings
from leonaid.application.crm import (
    CompanyData,
    CompanyRecord,
    CompanyUpdate,
    CrmGatewayError,
    PersonData,
    PersonRecord,
    PersonUpdate,
    PostalAddress,
)
from leonaid.application.sponsor_matching import (
    candidate_company_query as sponsor_candidate_company_query,
    normalize_match_name,
)

JsonObject = dict[str, Any]
ImportMode = Literal["dry-run", "apply"]
OPAQUE_ACTOR_ID = re.compile(r"^[A-Z][A-Z0-9]*(?:-[A-Z0-9]+)+$")


class ImportFailure(RuntimeError):
    """The import file or target system violated the import contract."""


class RowStatus(StrEnum):
    NEW = "new"
    UPDATE = "update"
    UNCHANGED = "unchanged"
    CONFLICT = "conflict"
    REJECTED = "rejected"


@dataclass(frozen=True, slots=True)
class ImportRow:
    row_number: int
    source_id: UUID | None
    record_type: str
    company_name: str | None
    given_name: str | None
    family_name: str | None
    email: str | None
    street_line_1: str | None
    postal_code: str | None
    city: str | None
    country: str | None
    validation_errors: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class RowPlan:
    row_number: int
    source_id: str | None
    record_type: str
    status: RowStatus
    message: str
    candidates: tuple[JsonObject, ...] = ()
    target_twenty_id: str | None = None
    applied: bool = False

    def to_json(self) -> JsonObject:
        payload = asdict(self)
        payload["error_code"] = f"ROW_{self.status.value.upper()}"
        return payload


@dataclass(frozen=True, slots=True)
class ImportResolution:
    decision: Literal["use-existing", "create-new"]
    target_twenty_id: UUID | None


def require_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise ImportFailure(f"erforderliche Umgebungsvariable fehlt: {name}")
    return value


def normalize_name(value: str) -> str:
    return normalize_match_name(value)


def candidate_company_query(value: str) -> str:
    return sponsor_candidate_company_query(value)


def optional_cell(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def load_mapping(path: Path) -> JsonObject:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ImportFailure(f"Mapping ist nicht lesbar: {path}") from error
    if not isinstance(value, dict) or value.get("schemaVersion") != 1:
        raise ImportFailure("Import-Mapping besitzt nicht schemaVersion 1")
    columns = value.get("columns")
    required = value.get("required")
    if not isinstance(columns, dict) or not isinstance(required, dict):
        raise ImportFailure("Import-Mapping enthält keine columns/required")
    return value


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_resolutions(path: Path | None) -> dict[str, ImportResolution]:
    if path is None:
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ImportFailure(f"Resolution-Datei ist nicht lesbar: {path}") from error
    if not isinstance(payload, dict) or payload.get("schemaVersion") != 1:
        raise ImportFailure("Resolution-Datei besitzt nicht schemaVersion 1")
    decisions = payload.get("decisions")
    if not isinstance(decisions, list):
        raise ImportFailure("Resolution-Datei enthält keine decisions-Liste")
    result: dict[str, ImportResolution] = {}
    for item in decisions:
        if not isinstance(item, dict):
            raise ImportFailure("Resolution-Entscheidung ist kein Objekt")
        source_id = item.get("sourceId")
        decision = item.get("decision")
        if not isinstance(source_id, str):
            raise ImportFailure("Resolution-Entscheidung enthält keine sourceId")
        try:
            normalized_source_id = str(UUID(source_id))
        except ValueError as error:
            raise ImportFailure("Resolution-sourceId ist keine UUID") from error
        if normalized_source_id in result:
            raise ImportFailure("Resolution-sourceId ist doppelt")
        if decision not in {"use-existing", "create-new"}:
            raise ImportFailure("Resolution-decision ist ungültig")
        decided_by = item.get("decidedBy")
        if (
            not isinstance(decided_by, str)
            or OPAQUE_ACTOR_ID.fullmatch(decided_by) is None
        ):
            raise ImportFailure("Resolution-decidedBy ist keine opake Actor-ID")
        decided_at = item.get("decidedAt")
        if not isinstance(decided_at, str):
            raise ImportFailure("Resolution-decidedAt fehlt")
        try:
            parsed_decision_time = datetime.fromisoformat(
                decided_at.replace("Z", "+00:00")
            )
        except ValueError as error:
            raise ImportFailure(
                "Resolution-decidedAt ist kein ISO-8601-Zeitpunkt"
            ) from error
        if parsed_decision_time.tzinfo is None:
            raise ImportFailure("Resolution-decidedAt benötigt eine Zeitzone")
        raw_target = item.get("targetTwentyId")
        target: UUID | None = None
        if decision == "use-existing":
            if not isinstance(raw_target, str):
                raise ImportFailure("use-existing benötigt eine targetTwentyId")
            try:
                target = UUID(raw_target)
            except ValueError as error:
                raise ImportFailure(
                    "Resolution-targetTwentyId ist keine UUID"
                ) from error
        elif raw_target is not None:
            raise ImportFailure("create-new darf keine targetTwentyId besitzen")
        result[normalized_source_id] = ImportResolution(
            decision=decision,
            target_twenty_id=target,
        )
    return result


def xlsx_rows(path: Path, sheet_name: str) -> tuple[list[str], list[list[object]]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as error:
        raise ImportFailure(f"Excel-Datei ist nicht lesbar: {path}") from error
    try:
        if sheet_name not in workbook.sheetnames:
            raise ImportFailure(
                f"Excel-Blatt fehlt: {sheet_name}; vorhanden: "
                + ", ".join(workbook.sheetnames)
            )
        worksheet = workbook[sheet_name]
        values: list[tuple[object, ...]] = []
        for row in worksheet.iter_rows():
            formulas = [
                cell.coordinate
                for cell in row
                if getattr(cell, "data_type", None) == "f"
            ]
            if formulas:
                raise ImportFailure(
                    "Formeln sind im Importblatt verboten: " + ", ".join(formulas)
                )
            values.append(tuple(cell.value for cell in row))
    finally:
        workbook.close()
    if not values:
        raise ImportFailure(f"Excel-Blatt ist leer: {sheet_name}")
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    rows = [list(row) for row in values[1:] if any(value is not None for value in row)]
    return headers, rows


def csv_rows(path: Path) -> tuple[list[str], list[list[object]]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            reader = csv.reader(handle, dialect)
            values = list(reader)
    except (OSError, csv.Error) as error:
        raise ImportFailure(f"CSV-Datei ist nicht lesbar: {path}") from error
    if not values:
        raise ImportFailure("CSV-Datei ist leer")
    return [value.strip() for value in values[0]], [
        cast(list[object], row)
        for row in values[1:]
        if any(value.strip() for value in row)
    ]


def load_rows(path: Path, sheet_name: str, mapping: JsonObject) -> list[ImportRow]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        headers, raw_rows = xlsx_rows(path, sheet_name)
    elif suffix in {".csv", ".tsv"}:
        headers, raw_rows = csv_rows(path)
    else:
        raise ImportFailure("unterstützte Formate sind .xlsx, .csv und .tsv")
    if len(headers) != len(set(headers)) or "" in headers:
        raise ImportFailure("Header müssen eindeutig und nicht leer sein")
    header_indexes = {name: index for index, name in enumerate(headers)}
    columns = cast(dict[str, str], mapping["columns"])
    missing_headers = sorted(set(columns.values()) - set(header_indexes))
    if missing_headers:
        raise ImportFailure("Pflichtspalten fehlen: " + ", ".join(missing_headers))

    result: list[ImportRow] = []
    source_counts: Counter[str] = Counter()
    for raw in raw_rows:
        values = {
            logical: optional_cell(
                raw[header_indexes[column]]
                if header_indexes[column] < len(raw)
                else None
            )
            for logical, column in columns.items()
        }
        raw_source_id = values["sourceId"]
        source_id: UUID | None = None
        errors: list[str] = []
        if raw_source_id is None:
            errors.append("source_id fehlt")
        else:
            try:
                source_id = UUID(raw_source_id)
            except ValueError:
                errors.append("source_id ist keine UUID")
            else:
                if source_id.version != 4:
                    errors.append("source_id muss UUIDv4 sein")
                source_counts[str(source_id)] += 1
        record_type = (values["recordType"] or "").upper()
        if record_type not in {"COMPANY", "PERSON"}:
            errors.append("record_type muss COMPANY oder PERSON sein")
        if record_type == "COMPANY" and values["companyName"] is None:
            errors.append("company_name fehlt für COMPANY")
        if record_type == "PERSON":
            if values["givenName"] is None:
                errors.append("given_name fehlt für PERSON")
            if values["familyName"] is None:
                errors.append("family_name fehlt für PERSON")
            if values["email"] is not None and (
                "@" not in values["email"]
                or values["email"].startswith("@")
                or values["email"].endswith("@")
            ):
                errors.append("email ist ungültig")
        result.append(
            ImportRow(
                row_number=len(result) + 2,
                source_id=source_id,
                record_type=record_type or "UNKNOWN",
                company_name=values["companyName"],
                given_name=values["givenName"],
                family_name=values["familyName"],
                email=values["email"],
                street_line_1=values["streetLine1"],
                postal_code=values["postalCode"],
                city=values["city"],
                country=values["country"],
                validation_errors=tuple(errors),
            )
        )

    duplicate_ids = {value for value, count in source_counts.items() if count > 1}
    if duplicate_ids:
        result = [
            (
                replace(
                    row,
                    validation_errors=(
                        *row.validation_errors,
                        "source_id ist innerhalb der Datei doppelt",
                    ),
                )
                if row.source_id is not None and str(row.source_id) in duplicate_ids
                else row
            )
            for row in result
        ]
    return result


def company_candidate(record: CompanyRecord) -> JsonObject:
    return {
        "twentyId": str(record.twenty_id),
        "name": record.data.name,
        "postalCode": record.data.address.postal_code,
        "city": record.data.address.city,
    }


def person_candidate(record: PersonRecord) -> JsonObject:
    return {
        "twentyId": str(record.twenty_id),
        "givenName": record.data.given_name,
        "familyName": record.data.family_name,
        "email": record.data.email,
        "companyTwentyId": (
            str(record.data.company_twenty_id)
            if record.data.company_twenty_id is not None
            else None
        ),
    }


def merged_address(row: ImportRow, existing: PostalAddress | None) -> PostalAddress:
    current = existing or PostalAddress()
    return PostalAddress(
        street_line_1=row.street_line_1 or current.street_line_1,
        street_line_2=current.street_line_2,
        postal_code=row.postal_code or current.postal_code,
        city=row.city or current.city,
        state=current.state,
        country=row.country or current.country,
    )


class ContactImporter:
    def __init__(
        self,
        gateway: TwentyCrmGateway,
        *,
        mode: ImportMode,
        resolutions: dict[str, ImportResolution] | None = None,
    ) -> None:
        self._gateway = gateway
        self._mode = mode
        self._resolutions = resolutions or {}

    async def process(self, row: ImportRow) -> RowPlan:
        if row.validation_errors:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id) if row.source_id else None,
                record_type=row.record_type,
                status=RowStatus.REJECTED,
                message="; ".join(row.validation_errors),
            )
        if row.source_id is None:
            raise AssertionError("validierte Zeile benötigt source_id")
        if row.record_type == "COMPANY":
            return await self._company(row)
        if row.record_type == "PERSON":
            return await self._person(row)
        raise AssertionError("validierte Zeile benötigt bekannten record_type")

    async def _company(self, row: ImportRow) -> RowPlan:
        assert row.source_id is not None
        assert row.company_name is not None
        correlation = f"crm-import:row:{row.row_number}:company"
        found = await self._gateway.search_companies(
            candidate_company_query(row.company_name),
            correlation_id=correlation,
        )
        exact_matches = tuple(
            record
            for record in found
            if normalize_name(record.data.name) == normalize_name(row.company_name)
        )
        resolution = self._resolutions.get(str(row.source_id))
        if resolution is not None and resolution.decision == "create-new":
            matches: tuple[CompanyRecord, ...] = ()
        elif resolution is not None:
            candidate_ids = {record.twenty_id for record in found}
            if resolution.target_twenty_id not in candidate_ids:
                raise ImportFailure(
                    f"Resolution für Zeile {row.row_number} verweist nicht auf "
                    "einen gemeldeten Firmenkandidaten"
                )
            selected = await self._gateway.get_company(
                resolution.target_twenty_id,
                correlation_id=f"{correlation}:resolution",
            )
            if selected is None:
                raise ImportFailure(
                    f"Resolution für Zeile {row.row_number} verweist auf "
                    "eine fehlende Firma"
                )
            matches = (selected,)
        else:
            matches = exact_matches
        if not matches and found and resolution is None:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.CONFLICT,
                message="ähnliche Firmen benötigen eine explizite Entscheidung",
                candidates=tuple(
                    company_candidate(record)
                    for record in sorted(found, key=lambda item: str(item.twenty_id))
                ),
            )
        if len(matches) > 1:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.CONFLICT,
                message="mehrere Firmen besitzen denselben normalisierten Namen",
                candidates=tuple(
                    company_candidate(record)
                    for record in sorted(
                        matches,
                        key=lambda item: str(item.twenty_id),
                    )
                ),
            )
        if not matches:
            if self._mode == "apply":
                created, receipt = await self._gateway.create_company(
                    row.source_id,
                    CompanyData(
                        name=row.company_name,
                        address=merged_address(row, None),
                    ),
                    correlation_id=correlation,
                )
                return RowPlan(
                    row_number=row.row_number,
                    source_id=str(row.source_id),
                    record_type=row.record_type,
                    status=RowStatus.NEW,
                    message="neue Firma angelegt",
                    target_twenty_id=str(receipt.twenty_id),
                    applied=True,
                )
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.NEW,
                message="neue Firma würde angelegt",
            )

        existing = matches[0]
        address = merged_address(row, existing.data.address)
        if address == existing.data.address:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.UNCHANGED,
                message="bestehende Firma ist bereits aktuell",
                candidates=(company_candidate(existing),),
                target_twenty_id=str(existing.twenty_id),
            )
        if self._mode == "apply":
            updated, receipt = await self._gateway.update_company(
                row.source_id,
                existing.twenty_id,
                CompanyUpdate(address=address),
                correlation_id=correlation,
            )
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.UPDATE,
                message="bestehende Firma kontrolliert aktualisiert",
                candidates=(company_candidate(updated),),
                target_twenty_id=str(receipt.twenty_id),
                applied=True,
            )
        return RowPlan(
            row_number=row.row_number,
            source_id=str(row.source_id),
            record_type=row.record_type,
            status=RowStatus.UPDATE,
            message="bestehende Firma würde kontrolliert aktualisiert",
            candidates=(company_candidate(existing),),
            target_twenty_id=str(existing.twenty_id),
        )

    async def _person(self, row: ImportRow) -> RowPlan:
        assert row.source_id is not None
        assert row.given_name is not None
        assert row.family_name is not None
        correlation = f"crm-import:row:{row.row_number}:person"
        company: CompanyRecord | None = None
        if row.company_name is not None:
            candidates = await self._gateway.search_companies(
                candidate_company_query(row.company_name),
                correlation_id=f"{correlation}:company",
            )
            company_matches = tuple(
                record
                for record in candidates
                if normalize_name(record.data.name) == normalize_name(row.company_name)
            )
            if len(company_matches) != 1:
                return RowPlan(
                    row_number=row.row_number,
                    source_id=str(row.source_id),
                    record_type=row.record_type,
                    status=(
                        RowStatus.CONFLICT if company_matches else RowStatus.REJECTED
                    ),
                    message=(
                        "mehrere passende Firmen für Person"
                        if company_matches
                        else "angegebene Firma für Person existiert nicht"
                    ),
                    candidates=tuple(
                        company_candidate(record) for record in company_matches
                    ),
                )
            company = company_matches[0]

        people = await self._gateway.search_people(
            given_name=row.given_name,
            family_name=row.family_name,
            correlation_id=correlation,
        )
        exact_matches = tuple(
            record
            for record in people
            if normalize_name(record.data.given_name) == normalize_name(row.given_name)
            and normalize_name(record.data.family_name)
            == normalize_name(row.family_name)
            and (company is None or record.data.company_twenty_id == company.twenty_id)
        )
        resolution = self._resolutions.get(str(row.source_id))
        if resolution is not None and resolution.decision == "create-new":
            matches: tuple[PersonRecord, ...] = ()
        elif resolution is not None:
            candidate_ids = {record.twenty_id for record in people}
            if resolution.target_twenty_id not in candidate_ids:
                raise ImportFailure(
                    f"Resolution für Zeile {row.row_number} verweist nicht auf "
                    "einen gemeldeten Personenkandidaten"
                )
            selected = await self._gateway.get_person(
                resolution.target_twenty_id,
                correlation_id=f"{correlation}:resolution",
            )
            if selected is None:
                raise ImportFailure(
                    f"Resolution für Zeile {row.row_number} verweist auf "
                    "eine fehlende Person"
                )
            matches = (selected,)
        else:
            matches = exact_matches
        if not matches and people and resolution is None:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.CONFLICT,
                message="ähnliche Personen benötigen eine explizite Entscheidung",
                candidates=tuple(
                    person_candidate(record)
                    for record in sorted(people, key=lambda item: str(item.twenty_id))
                ),
            )
        if len(matches) > 1:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.CONFLICT,
                message="mehrere Personen besitzen denselben Namensschlüssel",
                candidates=tuple(
                    person_candidate(record)
                    for record in sorted(
                        matches,
                        key=lambda item: str(item.twenty_id),
                    )
                ),
            )
        company_id = company.twenty_id if company else None
        if not matches:
            if self._mode == "apply":
                created, receipt = await self._gateway.create_person(
                    row.source_id,
                    PersonData(
                        given_name=row.given_name,
                        family_name=row.family_name,
                        email=row.email,
                        company_twenty_id=company_id,
                    ),
                    correlation_id=correlation,
                )
                return RowPlan(
                    row_number=row.row_number,
                    source_id=str(row.source_id),
                    record_type=row.record_type,
                    status=RowStatus.NEW,
                    message="neue Person angelegt",
                    candidates=(person_candidate(created),),
                    target_twenty_id=str(receipt.twenty_id),
                    applied=True,
                )
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.NEW,
                message="neue Person würde angelegt",
            )

        existing = matches[0]
        email_changed = row.email is not None and row.email != existing.data.email
        company_changed = (
            company_id is not None and company_id != existing.data.company_twenty_id
        )
        if not email_changed and not company_changed:
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.UNCHANGED,
                message="bestehende Person ist bereits aktuell",
                candidates=(person_candidate(existing),),
                target_twenty_id=str(existing.twenty_id),
            )
        update = PersonUpdate(
            email=row.email if email_changed else None,
            company_twenty_id=company_id if company_changed else None,
        )
        if self._mode == "apply":
            updated, receipt = await self._gateway.update_person(
                row.source_id,
                existing.twenty_id,
                update,
                correlation_id=correlation,
            )
            return RowPlan(
                row_number=row.row_number,
                source_id=str(row.source_id),
                record_type=row.record_type,
                status=RowStatus.UPDATE,
                message="bestehende Person kontrolliert aktualisiert",
                candidates=(person_candidate(updated),),
                target_twenty_id=str(receipt.twenty_id),
                applied=True,
            )
        return RowPlan(
            row_number=row.row_number,
            source_id=str(row.source_id),
            record_type=row.record_type,
            status=RowStatus.UPDATE,
            message="bestehende Person würde kontrolliert aktualisiert",
            candidates=(person_candidate(existing),),
            target_twenty_id=str(existing.twenty_id),
        )


def save_report(
    path: Path,
    *,
    source: Path,
    sheet: str,
    mode: ImportMode,
    plans: list[RowPlan],
    mapping: Path,
    resolutions: Path | None,
) -> None:
    counts = Counter(plan.status.value for plan in plans)
    report = {
        "schemaVersion": 1,
        "mode": mode,
        "source": source.name,
        "sourceSha256": sha256_file(source),
        "mappingSha256": sha256_file(mapping),
        "resolutionSha256": (
            sha256_file(resolutions) if resolutions is not None else None
        ),
        "sheet": sheet,
        "summary": {status.value: counts[status.value] for status in RowStatus},
        "appliedCount": sum(1 for plan in plans if plan.applied),
        "rows": [plan.to_json() for plan in plans],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    path.chmod(0o600)


async def run(arguments: argparse.Namespace) -> None:
    mapping_path = arguments.mapping.resolve()
    resolution_path = (
        arguments.resolutions.resolve() if arguments.resolutions is not None else None
    )
    mapping = load_mapping(mapping_path)
    resolutions = load_resolutions(resolution_path)
    sheet = arguments.sheet or str(mapping["defaultSheet"])
    source = arguments.source.resolve()
    rows = load_rows(source, sheet, mapping)
    row_source_ids = {str(row.source_id) for row in rows if row.source_id is not None}
    unused_resolutions = sorted(set(resolutions) - row_source_ids)
    if unused_resolutions:
        raise ImportFailure(
            "Resolution-Datei enthält Source-IDs außerhalb der Quelldatei"
        )
    settings = TwentyGatewaySettings(
        base_url=require_env("TWENTY_BASE_URL"),
        api_key=SecretStr(require_env("TWENTY_INTEGRATION_API_KEY")),
        timeout_seconds=5,
        page_size=20,
    )
    mode = cast(ImportMode, arguments.command)
    async with TwentyCrmGateway(settings) as gateway:
        importer = ContactImporter(
            gateway,
            mode=mode,
            resolutions=resolutions,
        )
        plans = [await importer.process(row) for row in rows]
    save_report(
        arguments.report.resolve(),
        source=source,
        sheet=sheet,
        mode=mode,
        plans=plans,
        mapping=mapping_path,
        resolutions=resolution_path,
    )
    counts = Counter(plan.status.value for plan in plans)
    print(
        "crm-import: OK: "
        + " ".join(f"{status.value}={counts[status.value]}" for status in RowStatus)
        + f" applied={sum(1 for plan in plans if plan.applied)} "
        + f"report={arguments.report}"
    )


def parser() -> argparse.ArgumentParser:
    result = argparse.ArgumentParser()
    subparsers = result.add_subparsers(dest="command", required=True)
    default_mapping = Path("infra/twenty/import-mapping.json")
    for command in ("dry-run", "apply"):
        command_parser = subparsers.add_parser(command)
        command_parser.add_argument("source", type=Path)
        command_parser.add_argument("--sheet")
        command_parser.add_argument("--mapping", type=Path, default=default_mapping)
        command_parser.add_argument("--resolutions", type=Path)
        command_parser.add_argument("--report", type=Path, required=True)
    return result


def main() -> int:
    arguments = parser().parse_args()
    try:
        asyncio.run(run(arguments))
    except (
        CrmGatewayError,
        ImportFailure,
        OSError,
        ValueError,
    ) as error:
        print(f"crm-import: ERROR: {error}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
